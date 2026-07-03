"""gerar_xlsx: layout da planilha, formatos e o vermelho de data alterada."""
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.services.remessas_export import gerar_xlsx
from app.storage.remessas_models import CicloRemessaRow, ConvenioRegistroRow


def _par(alterada=False):
    registro = ConvenioRegistroRow(id="r1", cod_empr=21, nome="Gov Espirito Santo",
                                   tipo_desconto="remessa", monitor_key=None,
                                   prod_beneficio=True, ativo=True)
    ciclo = CicloRemessaRow(
        id="c1", registro_id="r1", competencia="07/2026",
        competencia_inicio=date(2026, 7, 1),
        data_site=date(2026, 7, 26), data_site_alterada=alterada,
        data_envio=date(2026, 7, 23), valor_enviado=Decimal("1143.55"),
        qtd_contratos=3, beneficio_valor=Decimal("1143.55"), beneficio_qtd=3,
        observacao="ok", validado=True, corte_banksoft=date(2026, 7, 19),
    )
    return ciclo, registro


def _abrir(pares):
    return load_workbook(BytesIO(gerar_xlsx(pares, "07/2026")))


def test_layout_espelha_a_planilha():
    ws = _abrir([_par()]).active
    assert ws.title == "07-2026"
    assert [c.value for c in ws[1][:8]] == [
        "cod_empr", "Convenio", "competencia", "data_site", "data_envio",
        "valor_enviado", "qtd_contratos", "enviado"]
    linha = ws[2]
    assert linha[0].value == 21
    assert linha[1].value == "Gov Espirito Santo"
    assert linha[3].value.date() if hasattr(linha[3].value, "date") else linha[3].value  # data
    assert float(linha[5].value) == 1143.55
    assert linha[7].value == "SIM"          # enviado (tem data_envio)
    assert linha[9].value == "SIM"          # validado
    assert float(linha[12].value) == 1143.55  # benefício
    assert linha[5].number_format == "#,##0.00"
    assert linha[3].number_format == "dd/mm/yyyy"


def test_vermelho_quando_data_alterada():
    ws_alterada = _abrir([_par(alterada=True)]).active
    ws_normal = _abrir([_par(alterada=False)]).active
    assert ws_alterada.cell(row=2, column=4).fill.start_color.rgb.endswith("FFC7CE")
    assert not str(ws_normal.cell(row=2, column=4).fill.start_color.rgb).endswith("FFC7CE")


def test_automatico_sem_envio_rotulo():
    ciclo, registro = _par()
    registro.tipo_desconto = "automatico"
    ciclo.data_envio = None
    ws = _abrir([(ciclo, registro)]).active
    assert ws.cell(row=2, column=8).value == "AUTOMÁTICO"
