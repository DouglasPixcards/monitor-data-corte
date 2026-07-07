"""Importa competências do cortes.xlsx (planilha histórica da conciliação) para a plataforma.

Fluxo por competência (--competencia MM/YYYY):
- Convênio NOVO (cod_empr fora do cadastro): cria registro MANUAL (sem monitor_key; nome
  descoberto na própria planilha/Planilha4; produtos pelas colunas; tipo por heurística da
  observação) + ciclo completo da competência.
- Convênio JÁ cadastrado: MERGE conservador no ciclo — preenche SÓ campos vazios, nunca
  sobrescreve o que já está no painel. data_site só entra se vazia (o sync do monitor
  continua dono dela nos monitorados). Liga flags de produto no cadastro quando a planilha
  traz valores (senão as colunas ficariam invisíveis no grid).
- Regras acordadas: cods de TESTE ignorados (7, 13, 82); valor 0 importado LITERAL;
  'enviado' NÃO é importado (status é derivado) — divergências viram alerta; 'dp check'
  ignorada. Tudo auditado como usuário "importação histórica".

Sempre DRY-RUN por default; --apply grava. Idempotente (re-rodar não duplica nem re-preenche).

Uso:
    python scripts/importar_cortes_historico.py --arquivo cortes.xlsx --competencia 07/2026
    python scripts/importar_cortes_historico.py --arquivo cortes.xlsx --competencia 07/2026 --apply
"""
from __future__ import annotations

import argparse
import sys
import uuid
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from scripts.seed_convenios_registro import (  # noqa: E402
    _para_data, _para_decimal, _para_int, parece_automatico,
)

COD_TESTE = {7, 13, 82}          # convênios de teste — ignorar sempre
ABA_DADOS = "Cortes"
ABA_NOMES_EXTRA = "Planilha4"    # fonte adicional de nomes por cod_empr

# Campos do ciclo que participam do merge (planilha → modelo)
_CAMPOS_CICLO = (
    "data_envio", "valor_enviado", "qtd_contratos",
    "credito_valor", "credito_qtd", "beneficio_valor", "beneficio_qtd",
    "compras_valor", "compras_qtd", "corte_banksoft", "observacao", "validado",
)


def _comp_norm(v) -> str | None:
    """'07/26', '7/26', datetime(2026,7,1) → '07/2026'."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return f"{v.month:02d}/{v.year}"
    s = str(v).strip()
    if "/" in s:
        try:
            m, a = s.split("/")
            ano = int(a) + 2000 if len(a) <= 2 else int(a)
            return f"{int(m):02d}/{ano}"
        except ValueError:
            return None
    return None


def _preenchidos(d: dict) -> int:
    return sum(1 for v in d.values() if v not in (None, "", False))


def carregar(arquivo: Path, competencia: str):
    from openpyxl import load_workbook
    wb = load_workbook(arquivo, data_only=True)
    ws = wb[ABA_DADOS]
    cab = [str(c.value or "").strip() for c in ws[1]]
    i = {nome: cab.index(nome) for nome in (
        "cod_empr", "Convenio", "competencia", "data_site", "data_envio",
        "valor_enviado", "qtd_contratos", "enviado", "observacao", "validado",
        "Crédito", "Beneficio", "Compras", "data corte")}
    # Qtd por posição (coluna seguinte ao produto)
    qtd = {p: i[p] + 1 for p in ("Crédito", "Beneficio", "Compras")}

    nome_por_cod: dict[int, str] = {}
    linhas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = _para_int(row[i["cod_empr"]])
        if cod is None:
            continue
        if row[i["Convenio"]] not in (None, ""):
            nome_por_cod.setdefault(cod, str(row[i["Convenio"]]).strip())
        if _comp_norm(row[i["competencia"]]) != competencia:
            continue
        if cod in COD_TESTE:
            continue
        obs = str(row[i["observacao"]]).strip() if row[i["observacao"]] not in (None, "") else None
        linhas.append({
            "cod_empr": cod,
            "enviado_planilha": str(row[i["enviado"]] or "").strip().upper(),
            "obs": obs,
            "ciclo": {
                "data_site": _para_data(row[i["data_site"]]),
                "data_envio": _para_data(row[i["data_envio"]]),
                "valor_enviado": _para_decimal(row[i["valor_enviado"]]),   # 0 LITERAL
                "qtd_contratos": _para_int(row[i["qtd_contratos"]]),
                "credito_valor": _para_decimal(row[i["Crédito"]]),
                "credito_qtd": _para_int(row[qtd["Crédito"]]),
                "beneficio_valor": _para_decimal(row[i["Beneficio"]]),
                "beneficio_qtd": _para_int(row[qtd["Beneficio"]]),
                "compras_valor": _para_decimal(row[i["Compras"]]),
                "compras_qtd": _para_int(row[qtd["Compras"]]),
                "corte_banksoft": _para_data(row[i["data corte"]]),
                "observacao": obs,
                "validado": str(row[i["validado"]] or "").strip().upper() == "SIM",
            },
        })

    # nomes extras (Planilha4)
    if ABA_NOMES_EXTRA in wb.sheetnames:
        for row in wb[ABA_NOMES_EXTRA].iter_rows(min_row=2, values_only=True):
            cod = _para_int(row[0]) if row and row[0] is not None else None
            if cod is not None and len(row) > 1 and row[1] not in (None, ""):
                nome_por_cod.setdefault(cod, str(row[1]).strip())

    # dedup (cod, comp): mantém a linha mais completa
    por_cod: dict[int, dict] = {}
    duplicatas = []
    for l in linhas:
        atual = por_cod.get(l["cod_empr"])
        if atual is None:
            por_cod[l["cod_empr"]] = l
        else:
            duplicatas.append(l["cod_empr"])
            if _preenchidos(l["ciclo"]) > _preenchidos(atual["ciclo"]):
                por_cod[l["cod_empr"]] = l
    return list(por_cod.values()), nome_por_cod, duplicatas


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--arquivo", required=True, type=Path)
    parser.add_argument("--competencia", required=True, help="MM/YYYY (ex.: 07/2026)")
    parser.add_argument("--apply", action="store_true", help="Grava (default = dry-run)")
    args = parser.parse_args()

    from sqlalchemy import select

    from app.services.remessas_service import auditar, parse_competencia
    from app.storage import db
    from app.storage.remessas_models import CicloRemessaRow, ConvenioRegistroRow

    comp, inicio = parse_competencia(args.competencia)
    linhas, nome_por_cod, duplicatas = carregar(args.arquivo, comp)
    if not linhas:
        print(f"Nenhuma linha para a competência {comp} em {args.arquivo}.")
        return 1

    db.assert_ready()
    with db.session_scope() as session:
        registros = {r.cod_empr: r for r in session.execute(
            select(ConvenioRegistroRow)).scalars()}
        ciclos = {
            (c.registro_id, c.competencia): c for c in session.execute(
                select(CicloRemessaRow).where(CicloRemessaRow.competencia == comp)).scalars()
        }

        novos = [l for l in linhas if l["cod_empr"] not in registros]
        existentes = [l for l in linhas if l["cod_empr"] in registros]
        sem_nome = [l["cod_empr"] for l in novos if l["cod_empr"] not in nome_por_cod]
        sim_sem_data = [l["cod_empr"] for l in linhas
                        if l["enviado_planilha"] == "SIM" and l["ciclo"]["data_envio"] is None]

        print(f"Competência {comp} | linhas: {len(linhas)} (teste ignorado: {sorted(COD_TESTE)})")
        print(f"NOVOS (registro manual + ciclo): {len(novos)} | EXISTENTES (merge só-vazios): {len(existentes)}")
        if duplicatas:
            print(f"⚠️ duplicatas na planilha (mantida a mais completa): cods {sorted(set(duplicatas))}")
        if sem_nome:
            print(f"⚠️ novos SEM nome descoberto (entram como 'Convênio <cod>'): {sem_nome}")
        if sim_sem_data:
            print(f"⚠️ enviado=SIM sem data_envio (ficarão 'pendente'): {len(sim_sem_data)} cods {sorted(sim_sem_data)}")

        print(f"\n=== NOVOS ===")
        print(f"{'COD':>4} {'NOME':<40} {'TIPO':<10} {'PROD':<6} {'site':<11} {'envio':<11} {'valor':>12} {'banksoft'}")
        for l in sorted(novos, key=lambda x: x["cod_empr"]):
            c = l["ciclo"]
            tipo = "automatico" if parece_automatico(l["obs"]) else "remessa"
            prods = "".join(p[0].upper() for p, (v, q) in {
                "credito": (c["credito_valor"], c["credito_qtd"]),
                "beneficio": (c["beneficio_valor"], c["beneficio_qtd"]),
                "compras": (c["compras_valor"], c["compras_qtd"])}.items()
                if v is not None or q is not None) or "—"
            nome = nome_por_cod.get(l["cod_empr"], f"Convênio {l['cod_empr']}")
            print(f"{l['cod_empr']:>4} {nome[:40]:<40} "
                  f"{tipo:<10} {prods:<6} {str(c['data_site'] or '—'):<11} {str(c['data_envio'] or '—'):<11} "
                  f"{str(c['valor_enviado'] if c['valor_enviado'] is not None else '—'):>12} {c['corte_banksoft'] or '—'}")

        print(f"\n=== EXISTENTES (campos que o merge vai PREENCHER — vazios no painel) ===")
        plano_merge = []
        for l in sorted(existentes, key=lambda x: x["cod_empr"]):
            registro = registros[l["cod_empr"]]
            ciclo = ciclos.get((registro.id, comp))
            if ciclo is None:
                plano_merge.append((l, registro, None, list(_CAMPOS_CICLO) + ["data_site"]))
                print(f"{l['cod_empr']:>4} {registro.nome[:34]:<34} ciclo NÃO existe → será criado completo")
                continue
            enche = [campo for campo in _CAMPOS_CICLO
                     if getattr(ciclo, campo) in (None, False) and l["ciclo"].get(campo) not in (None, False)]
            if ciclo.data_site is None and l["ciclo"]["data_site"] is not None:
                enche.append("data_site")
            plano_merge.append((l, registro, ciclo, enche))
            print(f"{l['cod_empr']:>4} {registro.nome[:34]:<34} {', '.join(enche) if enche else 'nada a preencher'}")

        if not args.apply:
            print(f"\n[dry-run] Nada gravado. Rode com --apply para executar.")
            return 0

        # ── APPLY ────────────────────────────────────────────────────────────
        criados_reg = criados_ciclo = campos_preenchidos = flags_produto = 0
        for l in sorted(novos, key=lambda x: x["cod_empr"]):
            c = l["ciclo"]
            registro = ConvenioRegistroRow(
                id=str(uuid.uuid4()), cod_empr=l["cod_empr"],
                nome=nome_por_cod.get(l["cod_empr"], f"Convênio {l['cod_empr']}"),
                link_portal=None,
                tipo_desconto="automatico" if parece_automatico(l["obs"]) else "remessa",
                prod_credito=c["credito_valor"] is not None or c["credito_qtd"] is not None,
                prod_beneficio=c["beneficio_valor"] is not None or c["beneficio_qtd"] is not None,
                prod_compras=c["compras_valor"] is not None or c["compras_qtd"] is not None,
                monitor_key=None, ativo=True,
            )
            session.add(registro)
            auditar(session, entidade="registro", entidade_id=registro.id, acao="create",
                    usuario=None, valor_novo=f"importação histórica: {registro.cod_empr} {registro.nome}")
            session.flush()
            ciclo = CicloRemessaRow(
                id=str(uuid.uuid4()), registro_id=registro.id,
                competencia=comp, competencia_inicio=inicio,
                data_site=c["data_site"],
                data_site_origem="manual" if c["data_site"] else None,
                **{campo: c[campo] for campo in _CAMPOS_CICLO},
            )
            session.add(ciclo)
            auditar(session, entidade="ciclo", entidade_id=ciclo.id, acao="create",
                    usuario=None, valor_novo=f"importação histórica: {registro.nome} @ {comp}")
            criados_reg += 1
            criados_ciclo += 1

        for l, registro, ciclo, enche in plano_merge:
            c = l["ciclo"]
            if ciclo is None:
                ciclo = CicloRemessaRow(
                    id=str(uuid.uuid4()), registro_id=registro.id,
                    competencia=comp, competencia_inicio=inicio,
                    data_site=c["data_site"],
                    data_site_origem="manual" if c["data_site"] else None,
                    **{campo: c[campo] for campo in _CAMPOS_CICLO},
                )
                session.add(ciclo)
                auditar(session, entidade="ciclo", entidade_id=ciclo.id, acao="create",
                        usuario=None, valor_novo=f"importação histórica: {registro.nome} @ {comp}")
                criados_ciclo += 1
            else:
                for campo in enche:
                    anterior = getattr(ciclo, campo)
                    novo = c[campo]
                    setattr(ciclo, campo, novo)
                    if campo == "data_site":
                        ciclo.data_site_origem = ciclo.data_site_origem or "manual"
                    auditar(session, entidade="ciclo", entidade_id=ciclo.id, acao="update",
                            usuario=None, campo=campo, valor_anterior=anterior, valor_novo=novo)
                    campos_preenchidos += 1
            # liga flags de produto no cadastro quando a planilha traz valores
            for produto, (v, q) in {"credito": (c["credito_valor"], c["credito_qtd"]),
                                    "beneficio": (c["beneficio_valor"], c["beneficio_qtd"]),
                                    "compras": (c["compras_valor"], c["compras_qtd"])}.items():
                flag = f"prod_{produto}"
                if (v is not None or q is not None) and not getattr(registro, flag):
                    setattr(registro, flag, True)
                    auditar(session, entidade="registro", entidade_id=registro.id, acao="update",
                            usuario=None, campo=flag, valor_anterior=False, valor_novo=True)
                    flags_produto += 1

    print(f"\n✓ Importação {comp}: {criados_reg} registro(s) novo(s), {criados_ciclo} ciclo(s) criado(s), "
          f"{campos_preenchidos} campo(s) preenchido(s) no merge, {flags_produto} flag(s) de produto ligada(s).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
