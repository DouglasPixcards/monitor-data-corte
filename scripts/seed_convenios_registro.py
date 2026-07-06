"""Seed do cadastro de convênios (convenios_registro) + primeira carga de ciclos.

Dois layouts aceitos (detectados pelo cabeçalho):
- **Planilha BASE do painel** (tem coluna `monitor_key`): usa monitor_key/tipo_desconto
  EXPLÍCITOS e pré-carrega o ciclo da competência (--competencia) com data_envio, valores,
  qtds por produto, corte banksoft, validado e observação.
- **Planilha legada da conciliação** (sem `monitor_key`): auto-match do monitor_key por
  nome normalizado (difflib; ≥0.85 propõe, 0.60–0.85 pede revisão), 1:1 forçado.

Sempre DRY-RUN por default (tabela de revisão); nada é gravado sem --apply.
Idempotente: cod_empr já cadastrado é PULADO. Linha sem cod_empr é PULADA com aviso.

Uso:
    python scripts/seed_convenios_registro.py --planilha planilha-base-remessas.xlsx
    python scripts/seed_convenios_registro.py --planilha planilha-base-remessas.xlsx --apply --competencia 07/2026
"""
from __future__ import annotations

import argparse
import re
import sys
import unicodedata
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from difflib import SequenceMatcher
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

# ── Mapeamento da planilha LEGADA (sem monitor_key) ────────────────────────────
COLUNAS = {
    "cod_empr": "cod_empr",
    "nome": "Convenio",
    "observacao": "observacao",
    "credito": "Crédito",
    "beneficio": "Benefício",
    "compras": "Compras",
}
_MATCH_PROPOE = 0.85
_MATCH_REVISA = 0.60
_STOPWORDS = ("pref de ", "pref ", "prefeitura de ", "prefeitura ", "gov de ", "gov ",
              "governo de ", "governo do ", "governo da ", "municipio de ", "município de ")

# Notas geradas pela própria planilha-base — não são observação do usuário.
_OBS_GERADAS = ("sem coleta", "coleta antiga inválida", "datas divergem")


def normalizar_nome(s: str) -> str:
    s = unicodedata.normalize("NFKD", (s or "").lower())
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    for w in _STOPWORDS:
        s = s.replace(w, " ")
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def _variantes(nome_normalizado: str) -> list[str]:
    """O nome e o nome sem o sufixo de UF ('contagem mg' → também 'contagem')."""
    variantes = [nome_normalizado]
    tokens = nome_normalizado.split()
    if len(tokens) > 1 and len(tokens[-1]) == 2:  # UF no fim
        variantes.append(" ".join(tokens[:-1]))
    return variantes


def melhor_match(nome: str, candidatos: dict[str, str]) -> tuple[str | None, float]:
    """candidatos: monitor_key -> nome. Retorna (key, score) do melhor match por nome."""
    alvos = _variantes(normalizar_nome(nome))
    melhor, melhor_score = None, 0.0
    for key, cand_nome in candidatos.items():
        for cand in (normalizar_nome(cand_nome), normalizar_nome(key.replace("_", " "))):
            for alvo in alvos:
                score = SequenceMatcher(None, alvo, cand).ratio()
                if score > melhor_score:
                    melhor, melhor_score = key, score
    return melhor, round(melhor_score, 2)


def parece_automatico(observacao: str | None) -> bool:
    """Heurística FRACA (só sugestão pro dry-run): 'desc automatico', 'sem remessa'..."""
    o = normalizar_nome(observacao or "")
    return any(t in o for t in ("desc automatico", "desconto automatico", "sem remessa"))


# ── Parsing tolerante (Excel devolve datetime OU string em vários formatos) ────

def _para_data(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    m = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)     # D/M/YYYY
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", s)          # ISO
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _para_decimal(v) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v)).quantize(Decimal("0.01"))
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in s:                       # 1.234,56 → 1234.56
        s = s.replace(".", "").replace(",", ".")
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def _para_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).strip()))
    except ValueError:
        return None


def _para_bool(v) -> bool:
    return str(v).strip().lower() in ("sim", "true", "1", "x", "✓")


def _obs_do_usuario(v) -> str | None:
    s = (str(v).strip() if v is not None else "")
    if not s:
        return None
    low = s.lower()
    if any(low.startswith(g) for g in _OBS_GERADAS):
        return None   # nota gerada pela planilha-base, não é observação do usuário
    return s


# ── Leitura das planilhas ─────────────────────────────────────────────────────

def carregar_links(path: Path) -> dict[str, str]:
    """links_processadoras.xlsx → nome normalizado → URL (SÓ colunas CONVENIO/URL)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    cab = [str(c.value or "").strip().upper() for c in ws[1]]
    try:
        i_conv, i_url = cab.index("CONVENIO"), cab.index("URL")
    except ValueError:
        print(f"AVISO: {path} sem colunas CONVENIO/URL — links ignorados.")
        return {}
    return {
        normalizar_nome(str(row[i_conv])): str(row[i_url]).strip()
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[i_conv] and row[i_url]
    }


def carregar_planilha_base(path: Path) -> list[dict]:
    """Layout da planilha-base do painel (colunas explícitas; Qtd posicional após cada produto)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["remessas"] if "remessas" in wb.sheetnames else wb.active
    cab = [str(c.value or "").strip() for c in ws[1]]

    def idx(nome):
        return cab.index(nome) if nome in cab else None

    i = {n: idx(n) for n in ("cod_empr", "Convenio", "data_site", "data_envio",
                             "valor_enviado", "qtd_contratos", "observacao", "validado",
                             "corte banksoft", "monitor_key", "tipo_desconto")}
    # Qtd é duplicada — resolve por POSIÇÃO (coluna seguinte ao produto)
    prod_idx = {}
    for produto in ("Crédito", "Benefício", "Compras"):
        pi = idx(produto)
        prod_idx[produto] = (pi, pi + 1 if pi is not None else None)

    linhas, pulados = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def val(chave):
            j = i.get(chave)
            return row[j] if j is not None and j < len(row) else None

        nome = val("Convenio")
        if nome is None or str(nome).strip() == "":
            continue
        cod = _para_int(val("cod_empr"))
        if cod is None:
            pulados.append(str(nome).strip())
            continue

        def prod(nome_prod):
            vi, qi = prod_idx[nome_prod]
            v = _para_decimal(row[vi]) if vi is not None else None
            q = _para_int(row[qi]) if qi is not None else None
            return v, q

        cred_v, cred_q = prod("Crédito")
        ben_v, ben_q = prod("Benefício")
        com_v, com_q = prod("Compras")
        tipo = str(val("tipo_desconto") or "remessa").strip().lower()
        linhas.append({
            "cod_empr": cod,
            "nome": str(nome).strip(),
            "monitor_key": (str(val("monitor_key")).strip() or None) if val("monitor_key") else None,
            "tipo_desconto": tipo if tipo in ("remessa", "automatico") else "remessa",
            "prod_credito": cred_v is not None or cred_q is not None,
            "prod_beneficio": ben_v is not None or ben_q is not None,
            "prod_compras": com_v is not None or com_q is not None,
            "observacao": _obs_do_usuario(val("observacao")),
            "link_portal": None,
            # pré-carga do ciclo
            "ciclo": {
                "data_site": _para_data(val("data_site")),
                "data_envio": _para_data(val("data_envio")),
                "valor_enviado": _para_decimal(val("valor_enviado")),
                "qtd_contratos": _para_int(val("qtd_contratos")),
                "credito_valor": cred_v, "credito_qtd": cred_q,
                "beneficio_valor": ben_v, "beneficio_qtd": ben_q,
                "compras_valor": com_v, "compras_qtd": com_q,
                "corte_banksoft": _para_data(val("corte banksoft")),
                "validado": _para_bool(val("validado")),
            },
        })
    return linhas, pulados


def carregar_planilha_legada(path: Path) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    cab = {str(c.value or "").strip(): j for j, c in enumerate(ws[1])}
    faltando = [v for v in (COLUNAS["cod_empr"], COLUNAS["nome"]) if v not in cab]
    if faltando:
        raise SystemExit(f"ERRO: colunas obrigatórias ausentes: {faltando}. Visto: {list(cab)}")

    def val(row, chave):
        j = cab.get(COLUNAS[chave])
        return row[j] if j is not None and j < len(row) else None

    linhas, pulados = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod, nome = _para_int(val(row, "cod_empr")), val(row, "nome")
        if nome is None or str(nome).strip() == "":
            continue
        if cod is None:
            pulados.append(str(nome).strip())
            continue
        linhas.append({
            "cod_empr": cod, "nome": str(nome).strip(),
            "monitor_key": None, "tipo_desconto": None,   # resolvidos pelo auto-match
            "observacao": _obs_do_usuario(val(row, "observacao")),
            "prod_credito": val(row, "credito") is not None,
            "prod_beneficio": val(row, "beneficio") is not None,
            "prod_compras": val(row, "compras") is not None,
            "link_portal": None, "ciclo": None,
        })
    return linhas, pulados


def _tem_monitor_key(path: Path) -> bool:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb["remessas"] if "remessas" in wb.sheetnames else wb.active
    return any(str(c.value or "").strip() == "monitor_key" for c in ws[1])


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--planilha", required=True, type=Path)
    parser.add_argument("--links", type=Path, default=None)
    parser.add_argument("--apply", action="store_true", help="Grava (default = dry-run)")
    parser.add_argument("--competencia", default=None,
                        help="Pré-carrega os ciclos desta competência (MM/YYYY)")
    args = parser.parse_args()

    from app.core.loader import load_processadoras_config

    convenios_cfg = load_processadoras_config()["convenios"]
    monitor = {k: cfg.get("nome", k) for k, cfg in convenios_cfg.items()}
    modo_base = _tem_monitor_key(args.planilha)

    if modo_base:
        linhas, pulados = carregar_planilha_base(args.planilha)
    else:
        linhas, pulados = carregar_planilha_legada(args.planilha)

    links = carregar_links(args.links) if args.links else {}
    print(f"Layout: {'BASE (monitor_key explícito)' if modo_base else 'legado (auto-match)'} | "
          f"linhas: {len(linhas)} | puladas sem cod_empr: {len(pulados)} {pulados or ''}\n")

    print(f"{'COD':>4} {'NOME':<34} {'TIPO':<10} {'PROD':<6} {'MONITOR':<22} {'CICLO (envio/valor/banksoft)'}")
    print("-" * 112)
    usados: dict[str, int] = {}
    problemas = []
    for l in linhas:
        if modo_base:
            mk = l["monitor_key"]
            if mk and mk not in monitor:
                problemas.append(f"{l['nome']}: monitor_key desconhecida {mk!r}")
                l["monitor_key"] = None
            status = l["monitor_key"] or "manual"
        else:
            key, score = melhor_match(l["nome"], monitor)
            if score >= _MATCH_PROPOE:
                l["monitor_key"], status = key, f"{key} ({score}) ✓"
            elif score >= _MATCH_REVISA:
                l["monitor_key"], status = None, f"{key}? ({score}) REVISAR"
            else:
                l["monitor_key"], status = None, "manual"
            l["tipo_desconto"] = "automatico" if parece_automatico(l["observacao"]) else "remessa"
        if l["monitor_key"]:
            usados[l["monitor_key"]] = usados.get(l["monitor_key"], 0) + 1
        l["link_portal"] = links.get(normalizar_nome(l["nome"]))
        prods = "".join(p[0].upper() for p in ("credito", "beneficio", "compras")
                        if l[f"prod_{p}"]) or "—"
        c = l.get("ciclo") or {}
        resumo_ciclo = " / ".join(str(x) for x in (
            c.get("data_envio") or "—", c.get("valor_enviado") or "—",
            c.get("corte_banksoft") or "—"))
        print(f"{l['cod_empr']:>4} {l['nome'][:34]:<34} {l['tipo_desconto'] or '?':<10} "
              f"{prods:<6} {status:<22} {resumo_ciclo}")

    duplicados = {k for k, n in usados.items() if n > 1}
    if duplicados:
        print(f"\n⚠️ monitor_keys com MAIS de um cod_empr (1:1 violado — ninguém leva): {duplicados}")
        for l in linhas:
            if l["monitor_key"] in duplicados:
                l["monitor_key"] = None
    for p in problemas:
        print(f"⚠️ {p}")

    if not args.apply:
        print("\n[dry-run] Nada gravado. Revise e rode com --apply "
              "(+ --competencia MM/YYYY para pré-carregar os ciclos).")
        return 0

    from sqlalchemy import select

    from app.services.remessas_service import auditar, parse_competencia
    from app.storage import db
    from app.storage.remessas_models import CicloRemessaRow, ConvenioRegistroRow

    db.assert_ready()
    comp = inicio = None
    if args.competencia:
        comp, inicio = parse_competencia(args.competencia)

    criados = pulados_exist = ciclos_criados = 0
    with db.session_scope() as session:
        existentes = {c for (c,) in session.execute(select(ConvenioRegistroRow.cod_empr))}
        keys_usadas = {k for (k,) in session.execute(
            select(ConvenioRegistroRow.monitor_key)
            .where(ConvenioRegistroRow.monitor_key.is_not(None)))}
        for l in linhas:
            if l["cod_empr"] in existentes:
                pulados_exist += 1
                continue
            mk = l["monitor_key"] if l["monitor_key"] not in keys_usadas else None
            registro = ConvenioRegistroRow(
                id=str(uuid.uuid4()), cod_empr=l["cod_empr"], nome=l["nome"],
                link_portal=l["link_portal"], tipo_desconto=l["tipo_desconto"] or "remessa",
                prod_credito=l["prod_credito"], prod_beneficio=l["prod_beneficio"],
                prod_compras=l["prod_compras"], monitor_key=mk, ativo=True,
            )
            session.add(registro)
            if mk:
                keys_usadas.add(mk)
            auditar(session, entidade="registro", entidade_id=registro.id, acao="create",
                    usuario=None, valor_novo=f"seed: {l['cod_empr']} {l['nome']}")
            criados += 1
            # Sem relationship() entre os modelos o SQLAlchemy não ordena o INSERT
            # registro→ciclo sozinho — flush garante a FK antes do ciclo.
            session.flush()

            if comp is not None:
                c = l.get("ciclo") or {}
                ciclo = CicloRemessaRow(
                    id=str(uuid.uuid4()), registro_id=registro.id,
                    competencia=comp, competencia_inicio=inicio,
                    # data_site manual SÓ para não-monitorados; monitorados vêm do sync.
                    data_site=c.get("data_site") if mk is None else None,
                    data_site_origem="manual" if (mk is None and c.get("data_site")) else None,
                    data_envio=c.get("data_envio"),
                    valor_enviado=c.get("valor_enviado"),
                    qtd_contratos=c.get("qtd_contratos"),
                    credito_valor=c.get("credito_valor"), credito_qtd=c.get("credito_qtd"),
                    beneficio_valor=c.get("beneficio_valor"), beneficio_qtd=c.get("beneficio_qtd"),
                    compras_valor=c.get("compras_valor"), compras_qtd=c.get("compras_qtd"),
                    corte_banksoft=c.get("corte_banksoft"),
                    observacao=l["observacao"],
                    validado=bool(c.get("validado")),
                )
                session.add(ciclo)
                auditar(session, entidade="ciclo", entidade_id=ciclo.id, acao="create",
                        usuario=None, valor_novo=f"seed: {l['nome']} @ {comp}")
                ciclos_criados += 1

    print(f"\n✓ Seed: {criados} registro(s) criado(s), {pulados_exist} já existiam, "
          f"{ciclos_criados} ciclo(s) pré-carregado(s) em {comp or '—'}.")
    print("Próximo passo: rodar o sync (↻ no painel ou POST /remessas/sync) "
          "para puxar as data_site do monitor.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
