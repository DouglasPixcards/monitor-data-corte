"""Export .xlsx da competência — espelha o layout da planilha de conciliação.

Gera o workbook em memória (openpyxl). Datas como date (formato dd/mm/yyyy), valores
como número (#,##0.00), e data_site com fundo VERMELHO quando `data_site_alterada`
(a mesma convenção visual da planilha original).
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.remessas_service import status_ciclo

_CABECALHO = [
    "cod_empr", "Convenio", "competencia", "data_site", "data_envio",
    "valor_enviado", "qtd_contratos", "enviado", "observacao", "validado",
    "Crédito", "Qtd", "Benefício", "Qtd", "Compras", "Qtd", "corte banksoft",
]
_LARGURAS = [10, 38, 12, 12, 12, 14, 13, 11, 40, 10, 12, 8, 12, 8, 12, 8, 14]

_FMT_DATA = "dd/mm/yyyy"
_FMT_MOEDA = "#,##0.00"
_VERMELHO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_CAB_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_CAB_FONT = Font(bold=True, color="FFFFFF")

_STATUS_LABEL = {"automatico": "AUTOMÁTICO", "enviado": "SIM", "pendente": "NÃO"}


def gerar_xlsx(pares: list[tuple], competencia: str) -> bytes:
    """`pares` = [(CicloRemessaRow, ConvenioRegistroRow), ...] — retorna os bytes do .xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = competencia.replace("/", "-")

    ws.append(_CABECALHO)
    for col, largura in enumerate(_LARGURAS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura
        celula = ws.cell(row=1, column=col)
        celula.fill = _CAB_FILL
        celula.font = _CAB_FONT
    ws.freeze_panes = "A2"

    for linha, (ciclo, registro) in enumerate(pares, start=2):
        valores = [
            registro.cod_empr,
            registro.nome,
            ciclo.competencia,
            ciclo.data_site,
            ciclo.data_envio,
            ciclo.valor_enviado,
            ciclo.qtd_contratos,
            _STATUS_LABEL.get(status_ciclo(registro, ciclo), ""),
            ciclo.observacao,
            "SIM" if ciclo.validado else "",
            ciclo.credito_valor, ciclo.credito_qtd,
            ciclo.beneficio_valor, ciclo.beneficio_qtd,
            ciclo.compras_valor, ciclo.compras_qtd,
            ciclo.corte_banksoft,
        ]
        ws.append(valores)
        for col in (4, 5, 17):                    # datas
            ws.cell(row=linha, column=col).number_format = _FMT_DATA
        for col in (6, 11, 13, 15):               # moedas
            ws.cell(row=linha, column=col).number_format = _FMT_MOEDA
        if ciclo.data_site_alterada:              # o "vermelho" da planilha
            ws.cell(row=linha, column=4).fill = _VERMELHO

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
